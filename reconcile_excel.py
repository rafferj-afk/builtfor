"""
reconcile_excel.py

Reads all sports from SPORTS_DATA in builtfor.html, finds entries
that are missing from the Excel Profiles sheet, and appends them.

Run once to make the Excel the complete source of truth.
"""

import re, json
import openpyxl

HTML_PATH  = 'index.html'
EXCEL_PATH = 'sport_profile_database_v1_5.xlsx'

# ── 1. Parse SPORTS_DATA from HTML ────────────────────────────────────────────

with open(HTML_PATH, encoding='utf-8') as f:
    html = f.read()

start = html.index('const SPORTS_DATA = [')
depth, i = 0, start + len('const SPORTS_DATA = ')
end = i
while i < len(html):
    if html[i] == '[':   depth += 1
    elif html[i] == ']':
        depth -= 1
        if depth == 0:
            end = i + 1
            break
    i += 1

sports = json.loads(html[start + len('const SPORTS_DATA = '):end])
html_by_id = {s['id']: s for s in sports}
print(f'HTML sports parsed: {len(sports)}')

# ── 2. Get existing IDs from Excel ────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Profiles']

# Rows 1–2 are header rows; data starts at row 3
excel_ids = set()
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0]:
        excel_ids.add(str(row[0]))

print(f'Excel sports (existing): {len(excel_ids)}')

missing = [s for s in sports if s['id'] not in excel_ids]
print(f'Sports to append: {len(missing)}')
for s in missing:
    print(f'  {s["id"]:12s} {s["sport"]} ({s.get("sex","?")})')

# ── 3. Map HTML sport → Excel row ─────────────────────────────────────────────

def sport_to_row(s):
    p = s.get('physical', {})
    d = s.get('demands', {})
    c = s.get('categorical', {})
    ie = s.get('ireland', {})
    return (
        s.get('id'),
        s.get('sport'),
        s.get('position'),
        s.get('category'),
        p.get('heightMean'),
        p.get('heightSD'),
        p.get('weightMean'),
        p.get('weightSD'),
        p.get('bmiMean'),
        p.get('bfMean'),
        p.get('wingspanRatio'),
        p.get('sittingHeightRatio'),
        d.get('aerobic'),
        d.get('anaerobic'),
        d.get('maxStrength'),
        d.get('power'),
        d.get('speed'),
        d.get('endurance'),
        d.get('flexibility'),
        d.get('coordination'),
        d.get('reaction'),
        c.get('teamSize'),
        c.get('contact'),
        c.get('environment'),
        c.get('waterBased'),
        c.get('equipmentCost'),
        c.get('entryAge'),
        c.get('injuryRisk'),
        ie.get('accessibility'),
        ie.get('clubsPrevalent'),
        s.get('sources'),
        s.get('notes'),
        s.get('sex'),
    )

# ── 4. Append missing rows ────────────────────────────────────────────────────

for s in missing:
    ws.append(sport_to_row(s))

wb.save(EXCEL_PATH)
print(f'\nSaved. Excel Profiles sheet now has {ws.max_row - 2} sport rows.')
print('Next step: run export_sports.py to regenerate SPORTS_DATA in the HTML.')
